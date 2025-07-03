import webview
import sqlite3
import cv2
from pyzbar import pyzbar
from datetime import datetime
import os
import subprocess
import platform
import threading
import json
import base64
import time as time_module
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from contextlib import contextmanager
import shutil
import gc
import atexit


class QRScannerAPI:
    def __init__(self):
        self.data = []
        self.camera_thread = None
        self.scan_count = 0
        self.last_clear_date = datetime.now().date()
        self.camera_active = False
        self.cap = None
        self.scanned_codes = set()
        self._shutdown_event = threading.Event()
        self._window_closed = False
        self._ui_lock = threading.Lock()
        self._cleanup_done = False
        
        # Initialize settings before other components
        self.init_settings()
        self.init_db()
        self.load_today_records()
        self.start_midnight_checker()
        
        atexit.register(self.cleanup)
        self._camera_lock = threading.Lock()  # Add this line

    def init_settings(self):
        """Initialize settings system with default values"""
        self.settings_file = "_internal/data/settings.json"
        
        # Default settings
        self.default_settings = {
            "late_arrival_time": "08:15",
            "auto_save_interval": 300,  # seconds
            "camera_quality": 70,  # JPEG quality 1-100
            "camera_fps": 30,
            "duplicate_scan_timeout": 3,  # seconds
            "date_format": "%m/%d/%y",
            "time_format": "%H:%M:%S",
            "auto_backup": True,
            "backup_interval": 24,  # hours
            "sound_notifications": True,
            "visual_notifications": True,
            "auto_update_sf2": True,
            "camera_index": 0,
            "window_always_on_top": False,
            "dark_mode": False,
            "font_size": "medium"
        }
        
        # Load or create settings
        self.load_settings()

    def load_settings(self):
        """Load settings from file or create with defaults"""
        try:
            os.makedirs(os.path.dirname(self.settings_file), exist_ok=True)
            
            if os.path.exists(self.settings_file):
                with open(self.settings_file, 'r') as f:
                    saved_settings = json.load(f)
                
                # Merge with defaults to ensure all keys exist
                self.settings = {**self.default_settings, **saved_settings}
                
                # Validate settings
                self.validate_settings()
            else:
                self.settings = self.default_settings.copy()
                self.save_settings()
                
        except Exception as e:
            print(f"Error loading settings: {e}")
            self.settings = self.default_settings.copy()
            self.save_settings()

    def validate_settings(self):
        """Validate and fix invalid settings"""
        # Validate time format
        try:
            datetime.strptime(self.settings["late_arrival_time"], "%H:%M")
        except ValueError:
            self.settings["late_arrival_time"] = self.default_settings["late_arrival_time"]
        
        # Validate numeric ranges
        self.settings["camera_quality"] = max(1, min(100, self.settings["camera_quality"]))
        self.settings["camera_fps"] = max(1, min(60, self.settings["camera_fps"]))
        self.settings["duplicate_scan_timeout"] = max(1, min(30, self.settings["duplicate_scan_timeout"]))
        self.settings["auto_save_interval"] = max(60, min(3600, self.settings["auto_save_interval"]))
        self.settings["backup_interval"] = max(1, min(168, self.settings["backup_interval"]))
        self.settings["camera_index"] = max(0, min(10, self.settings["camera_index"]))

    def save_settings(self):
        """Save current settings to file"""
        try:
            os.makedirs(os.path.dirname(self.settings_file), exist_ok=True)
            with open(self.settings_file, 'w') as f:
                json.dump(self.settings, f, indent=2)
            return {'success': True, 'message': 'Settings saved successfully'}
        except Exception as e:
            return {'success': False, 'message': f'Error saving settings: {str(e)}'}

    def get_settings(self):
        """Get current settings for frontend"""
        return {
            'success': True,
            'settings': self.settings.copy()
        }

    def update_setting(self, key, value):
        """Update a single setting"""
        if key not in self.default_settings:
            return {'success': False, 'message': f'Unknown setting: {key}'}
        
        try:
            # Type conversion and validation
            if key == "late_arrival_time":
                # Validate time format
                datetime.strptime(value, "%H:%M")
            elif key in ["auto_save_interval", "camera_fps", "duplicate_scan_timeout", 
                        "backup_interval", "camera_index"]:
                value = int(value)
            elif key == "camera_quality":
                value = max(1, min(100, int(value)))
            elif key in ["auto_backup", "sound_notifications", "visual_notifications",
                        "auto_update_sf2", "window_always_on_top", "dark_mode"]:
                value = bool(value)
            
            self.settings[key] = value
            save_result = self.save_settings()
            
            if save_result['success']:
                # Apply certain settings immediately
                self.apply_setting_change(key, value)
                return {'success': True, 'message': f'Setting {key} updated successfully'}
            else:
                return save_result
                
        except Exception as e:
            return {'success': False, 'message': f'Invalid value for {key}: {str(e)}'}

    def update_settings(self, new_settings):
        """Update multiple settings at once"""
        try:
            # Backup current settings
            old_settings = self.settings.copy()
            
            # Validate and update each setting
            for key, value in new_settings.items():
                if key in self.default_settings:
                    result = self.update_setting(key, value)
                    if not result['success']:
                        # Restore old settings on any failure
                        self.settings = old_settings
                        return result
            
            return {'success': True, 'message': 'All settings updated successfully'}
            
        except Exception as e:
            return {'success': False, 'message': f'Error updating settings: {str(e)}'}

    def reset_settings(self):
        """Reset all settings to defaults"""
        try:
            self.settings = self.default_settings.copy()
            save_result = self.save_settings()
            
            if save_result['success']:
                # Apply all default settings
                for key, value in self.settings.items():
                    self.apply_setting_change(key, value)
                return {'success': True, 'message': 'Settings reset to defaults'}
            else:
                return save_result
                
        except Exception as e:
            return {'success': False, 'message': f'Error resetting settings: {str(e)}'}

    def apply_setting_change(self, key, value):
        """Apply setting changes that need immediate effect"""
        try:
            if key == "window_always_on_top":
                # This would need to be handled by the frontend
                self._safe_js_call(f'applyWindowSetting("always_on_top", {json.dumps(value)})')
            elif key == "dark_mode":
                self._safe_js_call(f'applyThemeSetting("dark_mode", {json.dumps(value)})')
            elif key == "font_size":
                self._safe_js_call(f'applyFontSetting("font_size", {json.dumps(value)})')
            elif key == "camera_index" and self.camera_active:
                # Restart camera with new index
                self.stop_camera()
                time_module.sleep(0.5)
                self.start_camera()
                
        except Exception as e:
            print(f"Error applying setting change for {key}: {e}")

    def export_settings(self):
        """Export settings to a file"""
        try:
            export_file = f"_internal/data/settings_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(export_file, 'w') as f:
                json.dump(self.settings, f, indent=2)
            
            return {
                'success': True, 
                'message': 'Settings exported successfully',
                'file': export_file
            }
        except Exception as e:
            return {'success': False, 'message': f'Error exporting settings: {str(e)}'}

    def import_settings(self, file_path):
        """Import settings from a file"""
        try:
            if not os.path.exists(file_path):
                return {'success': False, 'message': 'Settings file not found'}
            
            with open(file_path, 'r') as f:
                imported_settings = json.load(f)
            
            # Validate imported settings
            valid_settings = {}
            for key, value in imported_settings.items():
                if key in self.default_settings:
                    valid_settings[key] = value
            
            if valid_settings:
                result = self.update_settings(valid_settings)
                if result['success']:
                    return {
                        'success': True, 
                        'message': f'Imported {len(valid_settings)} settings successfully'
                    }
                else:
                    return result
            else:
                return {'success': False, 'message': 'No valid settings found in file'}
                
        except Exception as e:
            return {'success': False, 'message': f'Error importing settings: {str(e)}'}

    def set_window_closed(self):
        """Called when window is closing"""
        self._window_closed = True
        self._shutdown_event.set()
        threading.Thread(target=self.cleanup, daemon=True).start()
        return {'success': True}
    
    def load_today_records(self):
        """Load all today's attendance records from database"""
        today = datetime.now().strftime(self.settings["date_format"])
        self.cursor.execute("SELECT date, time, name FROM attendance WHERE date = ? ORDER BY time", (today,))
        rows = self.cursor.fetchall()
        self.data = [{'Date': row[0], 'Time': row[1], 'Name': row[2]} for row in rows]
        self.scan_count = len(self.data)
        
        self._safe_js_call(f'updateAttendanceTable({json.dumps(self.data)})')
    
    def record_attendance(self, name):
        """Record attendance for a person"""
        now = datetime.now()
        date_str = now.strftime(self.settings["date_format"])
        time_str = now.strftime(self.settings["time_format"])
        
        self.cursor.execute("SELECT 1 FROM attendance WHERE date = ? AND name = ?", (date_str, name))
        exists = self.cursor.fetchone()
        
        if exists:
            return {
                'success': False, 
                'message': f'{name} already scanned today',
                'type': 'duplicate'
            }
        
        try:
            self.cursor.execute("INSERT INTO attendance (date, time, name) VALUES (?, ?, ?)",
                              (date_str, time_str, name))
            self.conn.commit()
            
            self.load_today_records()
            
            return {
                'success': True,
                'message': f'Attendance of {name} recorded!',
                'type': 'success',
                'data': {'Date': date_str, 'Time': time_str, 'Name': name},
                'stats': self.get_stats()
            }
            
        except sqlite3.IntegrityError:
            return {
                'success': False,
                'message': f'{name} already recorded today',
                'type': 'duplicate'
            }
    
    def get_stats(self):
        """Get current statistics"""
        try:
            if not hasattr(self, 'data'):
                self.load_today_records()
                
            return {
                'scan_count': getattr(self, 'scan_count', 0),
                'camera_active': getattr(self, 'camera_active', False),
                'data': getattr(self, 'data', [])
            }
        except Exception as e:
            print(f"Error getting stats: {e}")
            return {
                'scan_count': 0,
                'camera_active': False,
                'data': []
            }

    def get_camera_status(self):
        """Get detailed camera status for debugging"""
        return {
            'camera_active': self.camera_active,
            'cap_exists': self.cap is not None,
            'cap_opened': self.cap.isOpened() if self.cap else False,
            'thread_alive': self.camera_thread.is_alive() if self.camera_thread else False,
            'shutdown_set': self._shutdown_event.is_set(),
            'window_closed': self._window_closed
        }

    def init_db(self):
        """Initialize SQLite database"""
        os.makedirs(os.path.join("_internal", "data"), exist_ok=True)
        self.conn = sqlite3.connect("_internal/data/attendance.db", check_same_thread=False)
        self.cursor = self.conn.cursor()
        
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS attendance (
                date TEXT,
                time TEXT,
                name TEXT,
                UNIQUE(date, name)
            )
        """)
        
        self.cursor.execute("""
            CREATE INDEX IF NOT EXISTS idx_date_name ON attendance(date, name)
        """)
        
        self.conn.commit()
    
    def start_midnight_checker(self):
        """Start background thread to check for new day"""
        def check_midnight():
            while not self._shutdown_event.is_set():
                try:
                    self._shutdown_event.wait(3600)
                    if self._shutdown_event.is_set():
                        break
                        
                    current_date = datetime.now().date()
                    if current_date != self.last_clear_date:
                        self.load_today_records()
                        self.last_clear_date = current_date
                        self._safe_js_call(f'handleNewDay({json.dumps(self.get_stats())})')
                except Exception as e:
                    print(f"Midnight checker error: {e}")
                    break
        
        threading.Thread(target=check_midnight, daemon=True).start()

    def check_camera_available(self):
        """Check if camera is available"""
        try:
            cap = cv2.VideoCapture(self.settings["camera_index"])
            if cap.isOpened():
                cap.release()
                return {'success': True, 'message': 'Camera available'}
            return {'success': False, 'message': 'Camera not available'}
        except Exception as e:
            return {'success': False, 'message': f'Camera error: {str(e)}'}
    
    def start_camera(self):
        """Start camera scanning"""
        if self.camera_active:
            return {'success': False, 'message': 'Camera already active'}
        
        if self._shutdown_event.is_set():
            return {'success': False, 'message': 'Application is shutting down'}
        
        try:
            self.cap = cv2.VideoCapture(self.settings["camera_index"])
            if not self.cap.isOpened():
                return {'success': False, 'message': 'Could not open camera'}
            
            self.camera_active = True
            self.scanned_codes.clear()
            
            self.camera_thread = threading.Thread(target=self._camera_loop, daemon=True)
            self.camera_thread.start()
            
            time_module.sleep(0.1)
            self._safe_js_call(f'updateCameraStatus(true)')
            
            return {'success': True, 'message': 'Camera started'}
        except Exception as e:
            self.camera_active = False
            return {'success': False, 'message': f'Camera error: {str(e)}'}
    
    def _camera_loop(self):
        """Main camera processing loop with improved error handling"""
        frame_count = 0
        fps_delay = 1.0 / self.settings["camera_fps"]
        
        # Only update status if window is still available
        if not self._window_closed:
            self._safe_js_call(f'updateCameraStatus(true)')
        
        while self.camera_active and self.cap and not self._shutdown_event.is_set() and not self._window_closed:
            if self._window_closed or self._shutdown_event.is_set():
                break
            try:
                ret, frame = self.cap.read()
                if not ret:
                    break
                
                barcodes = pyzbar.decode(frame)
                
                for barcode in barcodes:
                    if self._shutdown_event.is_set():
                        break
                        
                    qr_data = barcode.data.decode('utf-8')
                    
                    (x, y, w, h) = barcode.rect
                    cv2.rectangle(frame, (x, y), (x + w, y + h), (46, 204, 113), 3)
                    
                    text = f"Name: {qr_data}"
                    (text_w, text_h), _ = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
                    cv2.rectangle(frame, (x, y - text_h - 10), (x + text_w + 10, y), (46, 204, 113), -1)
                    cv2.putText(frame, text, (x + 5, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
                    
                    if qr_data not in self.scanned_codes:
                        self.scanned_codes.add(qr_data)
                        result = self.record_attendance(qr_data)
                        
                        if not self._shutdown_event.is_set():
                            self._safe_js_call(f'handleQRDetection({json.dumps(result)})')
                        
                        def remove_code():
                            if not self._shutdown_event.is_set():
                                self.scanned_codes.discard(qr_data)
                        
                        timer = threading.Timer(self.settings["duplicate_scan_timeout"], remove_code)
                        timer.daemon = True
                        timer.start()
                
                # Only update frame if window is still available
                frame_count += 1
                if frame_count % 2 == 0 and not self._window_closed:
                    try:
                        _, buffer = cv2.imencode('.jpg', frame, [cv2.IMWRITE_JPEG_QUALITY, self.settings["camera_quality"]])
                        frame_base64 = base64.b64encode(buffer).decode('utf-8')
                        success = self._safe_js_call(f'updateCameraFrame("data:image/jpeg;base64,{frame_base64}")')
                        if not success and not self._window_closed:
                            print("Failed to update camera frame - UI may be unavailable")
                    except Exception as e:
                        if not self._window_closed:
                            print(f"Frame encoding error: {e}")
                        break
                
                time_module.sleep(fps_delay)
                
            except Exception as e:
                if not self._shutdown_event.is_set() and not self._window_closed:
                    print(f"Camera loop error: {e}")
                break
        
        # Only update status if window is still available
        if not self._window_closed:
            self._safe_js_call(f'updateCameraStatus(false)')
        
        self._cleanup_camera()

    def _safe_js_call(self, js_code):
        """Safely execute JavaScript with proper error handling"""
        if self._window_closed or self._shutdown_event.is_set():
            return False
            
        try:
            if webview.windows and len(webview.windows) > 0:
                window = webview.windows[0]
                if hasattr(window, 'evaluate_js') and not self._window_closed:
                    window.evaluate_js(js_code)
                    return True
        except Exception as e:
            error_msg = str(e).lower()
            # Check for WebView2 disposal or other shutdown-related errors
            if any(keyword in error_msg for keyword in [
                "disposed", "cannot access", "objectdisposed", "webview2",
                "marshalinvoke", "invoke", "shutdown", "closed"
            ]):
                self._window_closed = True
                print(f"WebView disposed during JS call: {js_code[:50]}...")
                return False
            else:
                print(f"Unexpected JS call error: {e}")
                return False
        return False

    def _cleanup_camera(self):
        """Clean up camera resources"""
        if hasattr(self, '_camera_cleanup_done') and self._camera_cleanup_done:
            return
            
        print("Cleaning up camera resources...")
        self._camera_cleanup_done = True
        
        try:
            if self.cap:
                if self.cap.isOpened():
                    self.cap.release()
                self.cap = None
            
            # Give time for cleanup
            time_module.sleep(0.5)
            
            # Force cleanup of OpenCV windows
            try:
                cv2.destroyAllWindows()
                cv2.waitKey(1)
            except:
                pass
            
            # Force garbage collection
            gc.collect()
            
            print("Camera resources cleaned up")
        except Exception as e:
            print(f"Camera cleanup error: {e}")

    def cleanup(self):
        """Clean up resources properly"""
        if self._cleanup_done:
            return
            
        print("Starting cleanup...")
        self._cleanup_done = True
        self._window_closed = True
        self._shutdown_event.set()
        
        # Stop camera first
        self.camera_active = False
        
        # Wait for camera thread to finish
        if self.camera_thread and self.camera_thread.is_alive():
            self.camera_thread.join(timeout=2.0)
        
        # Clean up camera resources once
        if not hasattr(self, '_camera_cleanup_done'):
            self._cleanup_camera()
        
        # Close database connection
        try:
            if hasattr(self, 'conn') and self.conn:
                self.conn.close()
                print("Database connection closed")
        except Exception as e:
            print(f"Error closing database: {e}")
        
        print("Cleanup completed")

    def stop_camera(self):
        """Stop camera scanning"""
        self.camera_active = False
        
        # Wait for camera thread to finish
        if self.camera_thread and self.camera_thread.is_alive():
            self.camera_thread.join(timeout=2.0)  # Increased timeout
        
        # Ensure camera is properly released
        if self.cap:
            if self.cap.isOpened():
                self.cap.release()
            self.cap = None
        
        # Force OpenCV cleanup
        cv2.destroyAllWindows()
        cv2.waitKey(1)
        
        # Force garbage collection
        gc.collect()
        
        # Small delay to ensure cleanup
        time_module.sleep(0.5)
        
        self._safe_js_call(f'updateCameraStatus(false)')
        
        return {'success': True, 'message': 'Camera stopped'}

    def manual_entry(self, name):
        """Manual attendance entry"""
        if not name or not name.strip():
            return {'success': False, 'message': 'Name cannot be empty'}
        return self.record_attendance(name.strip())
    
    def toggle_camera(self):
        """Toggle camera on/off"""
        if self.camera_active:
            return self.stop_camera()
        else:
            return self.start_camera()
    
    @contextmanager
    def open_workbook(self, file_path):
        """Context manager for safely opening Excel workbooks"""
        wb = None
        try:
            if os.path.exists(file_path):
                os.chmod(file_path, 0o666)
            wb = load_workbook(file_path, data_only=False)
            yield wb
        except PermissionError as e:
            raise Exception(f"File is locked or in use: {e}")
        finally:
            if wb:
                try:
                    wb.close()
                except:
                    pass
                wb = None
            gc.collect()
            time_module.sleep(0.2)
    
    def update_sf2_automated(self):
        """Update SF2 Automated.xlsx with attendance data"""
        try:
            sf2_file = "_internal/data/SF2 Automated.xlsx"
            
            if not os.path.exists(sf2_file):
                return {'success': False, 'message': 'SF2 Automated.xlsx not found'}
            
            backup_file = f"{sf2_file}.backup"
            shutil.copy2(sf2_file, backup_file)
            
            with self.open_workbook(sf2_file) as wb:
                ws = wb.active
                
                dates = []
                for col in range(4, 29):
                    cell = ws.cell(row=11, column=col)
                    if cell.value:
                        try:
                            if isinstance(cell.value, datetime):
                                date_str = cell.value.strftime(self.settings["date_format"])
                            else:
                                date_str = datetime.strptime(str(cell.value), self.settings["date_format"]).strftime(self.settings["date_format"])
                            dates.append((col, date_str))
                        except ValueError:
                            continue
                
                all_dates = [date_str for _, date_str in dates]
                if not all_dates:
                    return {'success': False, 'message': 'No valid dates found in SF2'}
                
                placeholders = ','.join(['?' for _ in all_dates])
                self.cursor.execute(f"""
                    SELECT date, name, time FROM attendance 
                    WHERE date IN ({placeholders})
                """, all_dates)
                
                attendance_lookup = {}
                for date, name, time in self.cursor.fetchall():
                    attendance_lookup[(date, name)] = time
                
                changes_made = 0
                name_ranges = [(14, 43), (46, 75)]
                
                for start_row, end_row in name_ranges:
                    for row in range(start_row, end_row + 1):
                        name_cell = ws[f'B{row}']
                        if not name_cell.value:
                            continue
                        
                        name = str(name_cell.value).strip()
                        
                        for col, date_str in dates:
                            cell = ws.cell(row=row, column=col)
                            
                            if cell.data_type == 'f':
                                continue
                            
                            is_present = (date_str, name) in attendance_lookup
                            new_value = 0 if is_present else "x"
                            
                            if cell.value != new_value:
                                cell.value = new_value
                                if new_value == "x":
                                    cell.font = Font(color="000000")
                                changes_made += 1
                
                if changes_made > 0:
                    wb.save(sf2_file)
                    return {'success': True, 'message': f'Updated {changes_made} cells in SF2'}
                else:
                    return {'success': True, 'message': 'SF2 already up-to-date'}
                    
        except Exception as e:
            if os.path.exists(backup_file):
                shutil.copy2(backup_file, sf2_file)
            return {'success': False, 'message': f'Error updating SF2: {str(e)}'}
        finally:
            if os.path.exists(backup_file):
                try:
                    os.remove(backup_file)
                except:
                    pass
    
    def update_sf2_late_arrivals(self):
        """Update SF2 with black triangles for late arrivals"""
        try:
            sf2_file = "_internal/data/SF2 Automated.xlsx"
            
            if not os.path.exists(sf2_file):
                return {'success': False, 'message': 'SF2 Automated.xlsx not found'}
            
            backup_file = f"{sf2_file}.backup"
            shutil.copy2(sf2_file, backup_file)
            
            with self.open_workbook(sf2_file) as wb:
                ws = wb.active
                
                name_ranges = [
                    (14, 43),
                    (46, 75)
                ]
                
                dates = []
                for col in range(4, 29):
                    cell = ws.cell(row=11, column=col)
                    if cell.value:
                        try:
                            if isinstance(cell.value, datetime):
                                date_str = cell.value.strftime(self.settings["date_format"])
                            else:
                                date_str = datetime.strptime(str(cell.value), self.settings["date_format"]).strftime(self.settings["date_format"])
                            dates.append((col, date_str))
                        except ValueError:
                            continue
                
                late_arrivals_count = 0
                # Use the configurable late arrival time
                cutoff_time = datetime.strptime(self.settings["late_arrival_time"], "%H:%M").time()
                
                for start_row, end_row in name_ranges:
                    for row in range(start_row, end_row + 1):
                        name_cell = ws[f'B{row}']
                        if not name_cell.value:
                            continue
                        
                        name = str(name_cell.value).strip()
                        
                        for col, date_str in dates:
                            cell = ws.cell(row=row, column=col)
                            
                            if cell.data_type == 'f':
                                continue
                            
                            self.cursor.execute(
                                "SELECT time FROM attendance WHERE date=? AND name=?",
                                (date_str, name)
                            )
                            result = self.cursor.fetchone()
                            
                            if result:
                                try:
                                    attendance_time_str = result[0]
                                    attendance_time = datetime.strptime(attendance_time_str, self.settings["time_format"]).time()
                                    
                                    if attendance_time > cutoff_time:
                                        cell.value = None
                                        cell.number_format = 'General'
                                        
                                        self.add_late_marker_to_cell(ws, cell)
                                        late_arrivals_count += 1
                                        
                                except (ValueError, TypeError) as e:
                                    print(f"Error parsing time for {name} on {date_str}: {e}")
                                    continue
                
                if late_arrivals_count > 0:
                    wb.save(sf2_file)
                    return {'success': True, 'message': f'Added {late_arrivals_count} late arrival markers'}
                else:
                    return {'success': True, 'message': 'No late arrivals found'}
                    
        except Exception as e:
            if os.path.exists(backup_file):
                shutil.copy2(backup_file, sf2_file)
            return {'success': False, 'message': f'Error updating late arrivals: {str(e)}'}
        finally:
            if os.path.exists(backup_file):
                try:
                    os.remove(backup_file)
                except:
                    pass
    
    def add_late_marker_to_cell(self, worksheet, cell):
        """Add a black triangle image in the upper left corner of the cell"""
        try:
            triangle_image_path = "_internal/data/late.png"
            
            if not os.path.exists(triangle_image_path):
                print(f"Triangle image not found at {triangle_image_path}")
                self.add_triangle_border_fallback(cell)
                return
            
            img = Image(triangle_image_path)
            cell_coord = cell.coordinate
            img.anchor = cell_coord
            worksheet.add_image(img)
            
        except Exception as img_error:
            print(f"Error adding image: {img_error}")
            self.add_triangle_border_fallback(cell)
    
    def add_triangle_border_fallback(self, cell):
        """Fallback method to visually indicate late arrival with borders"""
        try:
            thin_border = Side(border_style="thin", color="000000")
            cell.border = Border(top=thin_border, left=thin_border)
            cell.comment = "Late arrival"
            
        except Exception as e:
            print(f"Error adding fallback border: {e}")
    
    def open_sf2_file(self):
        """Open SF2 file after updating"""
        try:
            # Update SF2 first
            update_result = self.update_sf2_automated()
            if not update_result['success']:
                return update_result
            
            # Update late arrivals
            late_result = self.update_sf2_late_arrivals()
            if not late_result['success']:
                return late_result
            
            # Then open the file
            sf2_file = "_internal/data/SF2 Automated.xlsx"
            file_path = os.path.abspath(sf2_file)
            if platform.system() == "Windows":
                os.startfile(file_path)
            elif platform.system() == "Darwin":
                subprocess.run(["open", file_path])
            else:
                subprocess.run(["xdg-open", file_path])
                
            return {'success': True, 'message': 'SF2 File Updated Successfully!'}
                
        except Exception as e:
            return {'success': False, 'message': f'Error opening SF2: {str(e)}'}



def on_window_close():
    """Handle window close event"""
    print("Window closing...")
    if hasattr(on_window_close, 'api'):
        on_window_close.api.set_window_closed()


def main():
    api = QRScannerAPI()
    on_window_close.api = api  # Store reference for cleanup
    
    html_file = os.path.join(os.path.dirname(__file__), 'data/index.html')

    
    window = webview.create_window(
        'SAM - School Attendance Management',
        html_file,
        js_api=api,
        width=900,
        height=700,
        min_size=(800, 600),
        resizable=True,
        maximized=True
    )
    
    window.events.closing += on_window_close
    
    try:
        webview.start(debug=False)
    except KeyboardInterrupt:
        print("Interrupted by user")
    except Exception as e:
        print(f"Application error: {e}")
    finally:
        print("Application shutting down...")
        api.cleanup()


if __name__ == "__main__":
    main()
